"""
Report export helpers for OptiATM.

Both generators take the SAME lightweight JSON-serializable `payload` dict --
a snapshot of whatever is currently on screen in the dashboard (KPIs, the
last optimization run, the last ATM-removal simulation, model metrics, and a
risk-analytics summary). All sections are optional except `kpis`; anything
not supplied is simply omitted from the report. This keeps the export
endpoints stateless -- the frontend already holds this data client-side
after its normal API calls, so we just format it.

Expected payload shape:
{
  "kpis": {
      "active_atms": int, "coverage_pct": str, "total_daily_tx": str, "avg_uptime": str
  },
  "optimization": {                      # optional
      "method": str, "k": int, "radius": float, "objective": str|None,
      "summary": str,
      "metrics": {coverage_before, coverage_after, coverage_increase,
                  avg_dist_before_km, avg_dist_after_km, avg_dist_reduction_km},
      "selected_candidates": [ {candidate_id, name, zone_name, site_type,
                                 latitude, longitude, rent_cost,
                                 predicted_daily_transactions, roi_index,
                                 payback_period}, ... ]
  },
  "removal_simulation": {...},           # optional, same shape as
                                          # ATMOptimizer.calculate_removal_impact()
  "model_metrics": {                     # optional
      "best_model_name": str, "r2": str, "mae": str, "rmse": str, "accuracy": str
  },
  "risk_summary": {                      # optional
      "risk_distribution": {"Low": int, "Medium": int, "High": int},
      "top_10_zones": {zone_name: {"risk_rate": float, "count": int}, ...}
  }
}
"""

import io
from datetime import datetime

BRAND_BLUE = "#2563eb"
BRAND_GREEN = "#16a34a"
BRAND_RED = "#dc2626"
BRAND_ORANGE = "#ea580c"


def _fmt(value, default="--"):
    return default if value is None or value == "" else value


# ---------------------------------------------------------------------------
# PDF
# ---------------------------------------------------------------------------

def create_network_map_drawing(selected_candidates=None):
    from reportlab.graphics.shapes import Drawing, Circle, Rect, String, Line
    from reportlab.lib import colors

    d = Drawing(480, 155)
    # Background card
    d.add(Rect(0, 0, 480, 155, fillColor=colors.HexColor("#f8fafc"), strokeColor=colors.HexColor("#cbd5e1"), strokeWidth=1, rx=4, ry=4))
    
    # Title overlay on drawing
    d.add(String(12, 138, "Hyderabad ATM Network Topology & Optimal Candidate Placements", fontSize=9, fontName="Helvetica-Bold", fillColor=colors.HexColor("#0f172a")))
    
    # Inner map boundary box
    d.add(Rect(12, 26, 456, 104, fillColor=colors.white, strokeColor=colors.HexColor("#e2e8f0"), strokeWidth=0.5))
    
    # Subtle background grid
    for x in range(60, 450, 60):
        d.add(Line(x, 26, x, 130, strokeColor=colors.HexColor("#f1f5f9"), strokeWidth=0.5))
    for y in range(45, 125, 25):
        d.add(Line(12, y, 468, y, strokeColor=colors.HexColor("#f1f5f9"), strokeWidth=0.5))
        
    # Sample background Hyderabad own ATMs scatter coordinates
    sample_coords = [
        (0.18, 0.38), (0.24, 0.58), (0.29, 0.32), (0.34, 0.72), (0.39, 0.48),
        (0.44, 0.28), (0.49, 0.62), (0.54, 0.38), (0.59, 0.68), (0.64, 0.22),
        (0.69, 0.52), (0.74, 0.42), (0.79, 0.58), (0.84, 0.32), (0.31, 0.50),
        (0.46, 0.76), (0.56, 0.20), (0.66, 0.80), (0.76, 0.26), (0.86, 0.56)
    ]
    
    for rx, ry in sample_coords:
        cx = 12 + rx * 456
        cy = 26 + ry * 104
        d.add(Circle(cx, cy, 3, fillColor=colors.HexColor("#2563eb"), strokeColor=colors.HexColor("#1d4ed8"), strokeWidth=0.5))
        
    # Plot selected candidates if provided
    selected_candidates = selected_candidates or []
    if selected_candidates:
        lats = [c.get("latitude", 17.41) for c in selected_candidates if isinstance(c.get("latitude"), (int, float))]
        lngs = [c.get("longitude", 78.44) for c in selected_candidates if isinstance(c.get("longitude"), (int, float))]
        
        min_lat, max_lat = (min(lats), max(lats)) if lats else (17.38, 17.48)
        min_lng, max_lng = (min(lngs), max(lngs)) if lngs else (78.38, 78.52)
        
        lat_span = (max_lat - min_lat) if max_lat > min_lat else 0.08
        lng_span = (max_lng - min_lng) if max_lng > min_lng else 0.08
        
        for idx, cand in enumerate(selected_candidates):
            lat = cand.get("latitude", 17.41)
            lng = cand.get("longitude", 78.44)
            norm_x = (lng - min_lng) / lng_span if lng_span > 0 else 0.5
            norm_y = (lat - min_lat) / lat_span if lat_span > 0 else 0.5
            
            cx = 32 + norm_x * 416
            cy = 38 + norm_y * 80
            
            cand_id = str(cand.get("candidate_id", f"C{idx+1}"))
            d.add(Circle(cx, cy, 6, fillColor=colors.HexColor("#16a34a"), strokeColor=colors.white, strokeWidth=1.5))
            d.add(String(cx + 8, cy - 3, cand_id, fontSize=8, fontName="Helvetica-Bold", fillColor=colors.HexColor("#0f172a")))
            
    # Bottom Legend bar
    d.add(Circle(20, 10, 3.5, fillColor=colors.HexColor("#2563eb"), strokeColor=colors.HexColor("#1d4ed8")))
    d.add(String(30, 7, "Active Own Network ATMs", fontSize=8, fontName="Helvetica", fillColor=colors.HexColor("#475569")))
    
    d.add(Circle(170, 10, 5, fillColor=colors.HexColor("#16a34a"), strokeColor=colors.white))
    d.add(String(182, 7, "Selected ATM Placements", fontSize=8, fontName="Helvetica-Bold", fillColor=colors.HexColor("#16a34a")))

    return d


def create_bar_chart_drawing(title, bars, width=468, value_suffix=""):
    """
    Simple horizontal bar-chart drawing for the PDF report, built directly
    with reportlab shapes (no external chart library needed) so it matches
    the same visual language as create_network_map_drawing.

    bars: list of (label, value, color_hex) tuples. Bar lengths are scaled
    relative to the largest |value| in the list.
    """
    from reportlab.graphics.shapes import Drawing, Rect, String
    from reportlab.lib import colors

    bar_h = 20
    gap = 10
    top_pad = 26
    bottom_pad = 8
    n = len(bars)
    height = top_pad + bottom_pad + n * (bar_h + gap) - gap

    d = Drawing(width, height)
    d.add(Rect(0, 0, width, height, fillColor=colors.HexColor("#f8fafc"), strokeColor=colors.HexColor("#cbd5e1"), strokeWidth=1, rx=4, ry=4))
    d.add(String(12, height - 18, title, fontSize=9, fontName="Helvetica-Bold", fillColor=colors.HexColor("#0f172a")))

    label_w = 150
    chart_left = 12 + label_w
    chart_right = width - 70
    chart_width = max(10, chart_right - chart_left)
    max_val = max([abs(v) for _, v, _ in bars] + [1.0])

    y = height - top_pad - bar_h
    for label, value, color_hex in bars:
        d.add(String(12, y + bar_h / 2 - 3, str(label), fontSize=8, fontName="Helvetica", fillColor=colors.HexColor("#334155")))
        bar_len = max(2, (abs(value) / max_val) * chart_width)
        d.add(Rect(chart_left, y + 4, bar_len, bar_h - 8, fillColor=colors.HexColor(color_hex), strokeColor=None))
        val_text = f"{value:,.1f}{value_suffix}" if isinstance(value, float) else f"{value:,}{value_suffix}"
        d.add(String(chart_left + bar_len + 6, y + bar_h / 2 - 3, val_text, fontSize=8, fontName="Helvetica-Bold", fillColor=colors.HexColor("#0f172a")))
        y -= (bar_h + gap)

    return d


def generate_pdf_report(payload: dict) -> io.BytesIO:
    from reportlab.lib.pagesizes import A4
    from reportlab.lib.units import mm
    from reportlab.lib import colors
    from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
    from reportlab.platypus import (
        SimpleDocTemplate, Paragraph, Spacer, Table, TableStyle, HRFlowable
    )

    buffer = io.BytesIO()
    doc = SimpleDocTemplate(
        buffer, pagesize=A4,
        topMargin=16 * mm, bottomMargin=16 * mm,
        leftMargin=16 * mm, rightMargin=16 * mm,
        title="OptiATM Network Report",
    )

    styles = getSampleStyleSheet()
    title_style = ParagraphStyle(
        "OptiTitle", parent=styles["Title"], textColor=colors.HexColor(BRAND_BLUE),
        fontSize=20, leading=22, spaceAfter=2, alignment=0
    )
    meta_style = ParagraphStyle(
        "OptiMeta", parent=styles["Normal"], textColor=colors.HexColor("#475569"),
        fontSize=9, spaceAfter=10,
    )
    h2_style = ParagraphStyle(
        "OptiH2", parent=styles["Heading2"], textColor=colors.HexColor("#0f172a"),
        fontSize=13, leading=16, spaceBefore=10, spaceAfter=6, fontName="Helvetica-Bold"
    )
    h3_style = ParagraphStyle(
        "OptiH3", parent=styles["Heading3"], textColor=colors.HexColor("#1e293b"),
        fontSize=10.5, leading=13, spaceBefore=8, spaceAfter=4, fontName="Helvetica-Bold"
    )
    body_style = ParagraphStyle(
        "OptiBody", parent=styles["Normal"], fontSize=9.5, leading=13.5,
        textColor=colors.HexColor("#334155"),
    )
    note_style = ParagraphStyle(
        "OptiNote", parent=styles["Normal"], fontSize=8.5, leading=11.5, textColor=colors.HexColor("#475569"),
    )
    tile_style = ParagraphStyle(
        "OptiTile", parent=styles["Normal"], alignment=1, leading=14,
    )

    def kv_table(rows, col_widths=None, colWidths=None):
        widths = col_widths or colWidths
        t = Table(rows, colWidths=widths, hAlign="LEFT")
        t.setStyle(TableStyle([
            ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor(BRAND_BLUE)),
            ("TEXTCOLOR", (0, 0), (-1, 0), colors.white),
            ("FONTNAME", (0, 0), (-1, 0), "Helvetica-Bold"),
            ("FONTSIZE", (0, 0), (-1, -1), 8.5),
            ("ROWBACKGROUNDS", (0, 1), (-1, -1), [colors.white, colors.HexColor("#f1f5f9")]),
            ("GRID", (0, 0), (-1, -1), 0.5, colors.HexColor("#cbd5e1")),
            ("VALIGN", (0, 0), (-1, -1), "MIDDLE"),
            ("TOPPADDING", (0, 0), (-1, -1), 4),
            ("BOTTOMPADDING", (0, 0), (-1, -1), 4),
            ("LEFTPADDING", (0, 0), (-1, -1), 6),
        ]))
        return t

    story = []

    # =========================================================================
    # PAGE 1 — Executive Summary
    # =========================================================================
    story.append(Paragraph("OptiATM &mdash; Network Report", title_style))
    story.append(Paragraph(
        f"Generated {datetime.now().strftime('%d %b %Y, %H:%M')} &middot; Hyderabad ATM Network",
        meta_style,
    ))
    story.append(HRFlowable(width="100%", color=colors.HexColor("#e2e8f0"), thickness=1))
    story.append(Spacer(1, 8))

    story.append(Paragraph("Executive Summary", h2_style))

    kpis = payload.get("kpis") or {}
    opt = payload.get("optimization") or {}
    selected = opt.get("selected_candidates") or []
    metrics = opt.get("metrics") or {}

    if selected:
        zones = list(dict.fromkeys([c.get("zone_name") for c in selected if c.get("zone_name")]))
        zone_str = ", ".join(zones[:3]) + (" and other key hubs" if len(zones) > 3 else "")
        cov_inc = _fmt(metrics.get("coverage_increase"), "0")
        dist_red = metrics.get("avg_dist_reduction_km", 0)
        dist_meters = f"{dist_red * 1000:.0f}m" if isinstance(dist_red, (int, float)) else "--"
        total_rent = sum([c.get("rent_cost", 0) for c in selected])
        paybacks = [c.get("payback_period") for c in selected if isinstance(c.get("payback_period"), (int, float)) and c.get("payback_period") > 0]
        avg_payback = (sum(paybacks) / len(paybacks)) if paybacks else 0

        exec_text = (
            f"We recommend opening <b>{len(selected)} new ATM(s)</b> in <b>{zone_str}</b>. "
            f"This raises area coverage by <b>+{cov_inc}%</b> and reduces average customer travel distance by <b>{dist_meters}</b>, "
            f"at a combined monthly rent of <b>INR {total_rent:,.0f}</b> with an estimated average payback period of <b>{avg_payback:.1f} months</b>."
        )

        k_req = opt.get("k")
        radius_km = opt.get("radius", 1.0)
        if isinstance(k_req, (int, float)) and len(selected) < k_req:
            exec_text += (
                f" Only <b>{len(selected)} of the requested {int(k_req)} sites</b> met the coverage threshold for selection; "
                f"the remaining candidates would not have added meaningful new coverage within the {radius_km}km radius."
            )
    else:
        exec_text = (
            f"This executive report synthesizes current network performance and optimization insights across Hyderabad. "
            f"The network currently comprises <b>{_fmt(kpis.get('active_atms'))} active ATMs</b> with <b>{_fmt(kpis.get('coverage_pct'))}</b> demographic coverage "
            f"and a total daily transaction volume of <b>{_fmt(kpis.get('total_daily_tx'))}</b>."
        )

    story.append(Paragraph(exec_text, body_style))
    story.append(Spacer(1, 10))

    # KPI Stat Tiles (2x2 grid)
    tile1 = Paragraph(f"<font size=16 color='{BRAND_BLUE}'><b>{_fmt(kpis.get('active_atms'))}</b></font><br/><font size=8 color='#475569'>Active Own ATMs</font>", tile_style)
    tile2 = Paragraph(f"<font size=16 color='{BRAND_BLUE}'><b>{_fmt(kpis.get('coverage_pct'))}</b></font><br/><font size=8 color='#475569'>Demographic Coverage</font>", tile_style)
    tile3 = Paragraph(f"<font size=16 color='{BRAND_BLUE}'><b>{_fmt(kpis.get('total_daily_tx'))}</b></font><br/><font size=8 color='#475569'>Total Daily Transaction Vol</font>", tile_style)
    tile4 = Paragraph(f"<font size=16 color='{BRAND_BLUE}'><b>{_fmt(kpis.get('avg_uptime'))}</b></font><br/><font size=8 color='#475569'>Avg Network Uptime (30-day)</font>", tile_style)

    tile_table = Table([
        [tile1, tile2],
        [tile3, tile4]
    ], colWidths=[234, 234], hAlign="LEFT")
    tile_table.setStyle(TableStyle([
        ("BACKGROUND", (0, 0), (-1, -1), colors.HexColor("#f8fafc")),
        ("GRID", (0, 0), (-1, -1), 0.5, colors.HexColor("#cbd5e1")),
        ("VALIGN", (0, 0), (-1, -1), "MIDDLE"),
        ("TOPPADDING", (0, 0), (-1, -1), 8),
        ("BOTTOMPADDING", (0, 0), (-1, -1), 8),
    ]))
    story.append(tile_table)
    story.append(Spacer(1, 12))

    # "How These Recommendations Were Generated" -- plain-language trust box
    # so a bank customer understands the approach without needing technical
    # background. Reflects the two-stage ML pre-filter (see app.py
    # apply_ml_prefilter) when that's what actually ran for this result.
    prefilter = opt.get("ml_prefilter")
    methodology_text = (
        "Candidate ATM sites are scored by a machine-learning model trained on local foot traffic, population "
        "density, income levels, and competitor proximity across the existing network, combined with a risk "
        "model that flags each site's likelihood of underperforming."
    )
    if prefilter:
        methodology_text += (
            f" For this recommendation, the {prefilter.get('total_candidates', '--')} candidate sites were first "
            f"narrowed to the {prefilter.get('shortlisted_candidates', '--')} strongest performers by predicted value, "
            f"then a spatial optimization algorithm selected the final locations from that shortlist to maximize "
            f"real-world population coverage and minimize customer travel distance."
        )
    else:
        methodology_text += (
            " A spatial optimization algorithm then selects the final locations to maximize real-world coverage, "
            "minimize travel distance, or maximize revenue -- depending on the strategy chosen for this run."
        )
    method_box = Table([[Paragraph(f"<b>How These Recommendations Were Generated:</b> {methodology_text}", note_style)]], colWidths=[468], hAlign="LEFT")
    method_box.setStyle(TableStyle([
        ("BACKGROUND", (0, 0), (-1, -1), colors.HexColor("#eff6ff")),
        ("GRID", (0, 0), (-1, -1), 0.5, colors.HexColor("#bfdbfe")),
        ("TOPPADDING", (0, 0), (-1, -1), 8),
        ("BOTTOMPADDING", (0, 0), (-1, -1), 8),
        ("LEFTPADDING", (0, 0), (-1, -1), 10),
        ("RIGHTPADDING", (0, 0), (-1, -1), 10),
    ]))
    story.append(method_box)
    story.append(Spacer(1, 12))

    # Map graphic
    story.append(create_network_map_drawing(selected))

    # =========================================================================
    # PAGE 2 — Why These Sites
    # =========================================================================
    if opt and selected:
        story.append(Spacer(1, 16))
        story.append(Paragraph("Why These Sites (Recommended Placements)", h2_style))
        story.append(Paragraph(
            "Each candidate location was selected by evaluating local demographic density, competitor proximity, "
            "predicted transaction volume, and financial payback period.",
            body_style
        ))
        story.append(Spacer(1, 10))

        for c in selected:
            payback = c.get("payback_period")
            payback_text = f"{payback:.1f} months" if isinstance(payback, (int, float)) and payback > 0 else "Never"
            
            reasons = c.get("top_reasons") or []
            if isinstance(reasons, list) and len(reasons) > 0:
                reasons_str = " &bull; ".join(reasons)
            else:
                reasons_str = "High foot traffic corridor &amp; strong demographic demand coverage"

            card_rows = [
                [
                    Paragraph(f"<b>Site ID: {c.get('candidate_id', '--')} &mdash; {c.get('name', 'Candidate')}</b>", ParagraphStyle("cHead", parent=body_style, textColor=colors.HexColor(BRAND_BLUE), fontName="Helvetica-Bold", fontSize=10)),
                    Paragraph(f"<b>Zone:</b> {c.get('zone_name', '--')}", ParagraphStyle("cSub", parent=body_style, fontSize=9)),
                    Paragraph(f"<b>Site Type:</b> {c.get('site_type', '--')}", ParagraphStyle("cSub2", parent=body_style, fontSize=9))
                ],
                [
                    Paragraph(f"<font size=7.5 color='#64748b'>PREDICTED DAILY TX</font><br/><b>{c.get('predicted_daily_transactions', 0):,.0f} tx/day</b>", body_style),
                    Paragraph(f"<font size=7.5 color='#64748b'>MONTHLY RENT</font><br/><b>INR {c.get('rent_cost', 0):,.0f}</b>", body_style),
                    Paragraph(f"<font size=7.5 color='#64748b'>ROI / PAYBACK</font><br/><b>ROI {c.get('roi_index', 0):.2f} ({payback_text})</b>", body_style)
                ],
                [
                    Paragraph(f"<font color='#0f172a'><b>Key Drivers:</b></font> {reasons_str}", ParagraphStyle("cDrivers", parent=body_style, fontSize=8.5, leading=11.5, textColor=colors.HexColor("#334155"))),
                    "", ""
                ]
            ]

            card = Table(card_rows, colWidths=[200, 134, 134], hAlign="LEFT")
            card.setStyle(TableStyle([
                ("SPAN", (0, 2), (2, 2)),
                ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#f1f5f9")),
                ("BACKGROUND", (0, 1), (-1, -1), colors.white),
                ("GRID", (0, 0), (-1, -1), 0.5, colors.HexColor("#cbd5e1")),
                ("VALIGN", (0, 0), (-1, -1), "MIDDLE"),
                ("TOPPADDING", (0, 0), (-1, -1), 6),
                ("BOTTOMPADDING", (0, 0), (-1, -1), 6),
                ("LEFTPADDING", (0, 0), (-1, -1), 8),
                ("RIGHTPADDING", (0, 0), (-1, -1), 8),
            ]))
            story.append(card)
            story.append(Spacer(1, 8))

    # =========================================================================
    # PAGE 3 — Network Impact & Financials
    # =========================================================================
    if opt:
        story.append(Spacer(1, 16))
        story.append(Paragraph("Network Impact & Financial Rollup", h2_style))
        story.append(Paragraph("Population Coverage & Customer Travel Distance Impact", h3_style))
        
        m = opt.get("metrics") or {}
        cov_before = m.get("coverage_before")
        cov_after = m.get("coverage_after")
        if isinstance(cov_before, (int, float)) and isinstance(cov_after, (int, float)):
            story.append(create_bar_chart_drawing(
                "Population Coverage: Before vs. After Optimization",
                [
                    ("Before Optimization", float(cov_before), "#94a3b8"),
                    ("After Optimization", float(cov_after), BRAND_BLUE),
                ],
                value_suffix="%"
            ))
            story.append(Spacer(1, 8))
        story.append(kv_table([
            ["Impact Metric", "Before Optimization", "After Optimization", "Net Change"],
            ["Population Coverage (%)", f"{_fmt(m.get('coverage_before'))}%", f"{_fmt(m.get('coverage_after'))}%", f"+{_fmt(m.get('coverage_increase'))}%"],
            ["Avg Travel Distance (km)", f"{_fmt(m.get('avg_dist_before_km'))}", f"{_fmt(m.get('avg_dist_after_km'))}", f"-{_fmt(m.get('avg_dist_reduction_km'))}"],
        ], colWidths=[180, 100, 100, 88]))

        if selected:
            story.append(Spacer(1, 10))
            story.append(Paragraph("Financial Rollup & Investment Summary", h3_style))
            total_rent = sum([c.get("rent_cost", 0) for c in selected])
            total_pred_tx = sum([c.get("predicted_daily_transactions", 0) for c in selected])
            # Assuming standard Interchange fee estimate of ~INR 15.35/tx and 30 days/month
            total_est_revenue = total_pred_tx * 15.35 * 30
            paybacks = [c.get("payback_period") for c in selected if isinstance(c.get("payback_period"), (int, float)) and c.get("payback_period") > 0]
            avg_payback_val = (sum(paybacks) / len(paybacks)) if paybacks else 0

            story.append(kv_table([
                ["Financial Rollup Metric", "Combined Value"],
                ["Total Monthly Rent Commitment", f"INR {total_rent:,.0f}"],
                ["Combined Predicted Daily Transactions", f"{total_pred_tx:,.0f} tx/day"],
                ["Estimated Monthly Network Revenue", f"INR {total_est_revenue:,.0f}"],
                ["Combined Average Payback Period", f"{avg_payback_val:.1f} months"],
            ], colWidths=[260, 208]))

        # Removal simulation block if present
        rem = payload.get("removal_simulation")
        if rem:
            story.append(Spacer(1, 12))
            story.append(Paragraph("ATM Removal Impact Simulation", h3_style))
            story.append(Paragraph(
                f"<b>Target Site:</b> {_fmt(rem.get('atm_id'))} ({_fmt(rem.get('zone_name'))}, {_fmt(rem.get('site_type'))}) &nbsp;&nbsp; "
                f"<b>Service Radius:</b> {_fmt(rem.get('radius_km'))} km",
                body_style,
            ))
            story.append(Spacer(1, 4))
            story.append(kv_table([
                ["Impact Metric", "Before Removal", "After Removal", "Net Loss/Increase"],
                ["Population Coverage (%)", f"{_fmt(rem.get('coverage_before'))}%", f"{_fmt(rem.get('coverage_after'))}%", f"-{_fmt(rem.get('coverage_loss'))}%"],
                ["Avg Travel Distance (km)", f"{_fmt(rem.get('avg_dist_before_km'))}", f"{_fmt(rem.get('avg_dist_after_km'))}", f"+{_fmt(rem.get('avg_dist_increase_km'))}"],
            ], colWidths=[180, 100, 100, 88]))
            story.append(Spacer(1, 4))
            story.append(kv_table([
                ["Risk Factor / Financial Impact", "Value"],
                ["Population Losing All Coverage", f"{_fmt(rem.get('affected_population')):,}" if isinstance(rem.get('affected_population'), int) else str(_fmt(rem.get('affected_population')))],
                ["Avg Daily Transactions at Risk", str(_fmt(rem.get('avg_daily_transactions')))],
                ["Monthly Revenue at Risk (INR)", f"{rem.get('monthly_revenue_at_risk', 0):,.0f}" if rem.get('monthly_revenue_at_risk') is not None else "--"],
                ["Monthly Rent Saved (INR)", f"{rem.get('monthly_rent_saved', 0):,.0f}" if rem.get('monthly_rent_saved') is not None else "--"],
                ["Net Monthly Profit Impact (INR)", f"{rem.get('net_monthly_profit_impact', 0):,.0f}" if rem.get('net_monthly_profit_impact') is not None else "--"],
            ], colWidths=[260, 208]))

    # =========================================================================
    # PAGE 4 — Model Confidence
    # =========================================================================
    mm_data = payload.get("model_metrics")
    if mm_data:
        story.append(Spacer(1, 16))
        story.append(Paragraph("Model Confidence & Evaluation", h2_style))

        r2_raw = mm_data.get("r2")
        mae_raw = mm_data.get("mae")

        try:
            r2_val = float(r2_raw) if r2_raw is not None else None
            mae_val = float(mae_raw) if mae_raw is not None else None
        except (ValueError, TypeError):
            r2_val = None
            mae_val = None

        if r2_val is not None and r2_val > 0 and mae_val is not None:
            confidence_text = (
                f"Our machine learning demand prediction model explains <b>{r2_val * 100:.0f}%</b> of the variation in ATM "
                f"transaction volume across Hyderabad, with an average prediction error of <b>~{mae_val:.0f} transactions/day</b> "
                f"on held-out test data."
            )
        else:
            confidence_text = (
                "Our machine learning demand prediction pipeline evaluates candidate locations using models trained on "
                "historical transaction logs, foot traffic metrics, and demographic density across the network."
            )

        story.append(Paragraph(confidence_text, body_style))
        story.append(Spacer(1, 10))

        story.append(kv_table([
            ["Model Performance Metric", "Value"],
            ["Active Selected Model", str(_fmt(mm_data.get("best_model_name")))],
            ["R\u00b2 Score (Held-Out Test Set)", str(_fmt(mm_data.get("r2")))],
            ["MAE (Transactions / Day)", str(_fmt(mm_data.get("mae")))],
            ["RMSE (Transactions / Day)", str(_fmt(mm_data.get("rmse")))],
            ["Prediction Accuracy (%)", str(_fmt(mm_data.get("accuracy")))],
        ], colWidths=[260, 208]))

        story.append(Spacer(1, 16))

        # Legal Disclaimer Box
        disc_table = Table([[
            Paragraph(
                "<b>Notice &amp; Disclaimer:</b> Figures are model estimates from OptiATM's synthetic demand-prediction "
                "and optimization pipeline, intended to support strategic bank planning decisions rather than serve as guaranteed financial outcomes.",
                note_style
            )
        ]], colWidths=[468], hAlign="LEFT")
        disc_table.setStyle(TableStyle([
            ("BACKGROUND", (0, 0), (-1, -1), colors.HexColor("#f1f5f9")),
            ("GRID", (0, 0), (-1, -1), 0.5, colors.HexColor("#cbd5e1")),
            ("TOPPADDING", (0, 0), (-1, -1), 8),
            ("BOTTOMPADDING", (0, 0), (-1, -1), 8),
            ("LEFTPADDING", (0, 0), (-1, -1), 10),
            ("RIGHTPADDING", (0, 0), (-1, -1), 10),
        ]))
        story.append(disc_table)

    # =========================================================================
    # PAGE 5 — Risk Flags
    # =========================================================================
    risk = payload.get("risk_summary")
    if risk:
        story.append(Spacer(1, 16))
        story.append(Paragraph("Risk Analytics & Watch List", h2_style))
        dist = risk.get("risk_distribution") or {}
        high_count = dist.get("High", 0)
        story.append(Paragraph(
            f"<b>{high_count} existing ATMs</b> are currently on the risk watch list requiring operational or site review.",
            body_style
        ))
        story.append(Spacer(1, 8))

        low_n = dist.get("Low")
        med_n = dist.get("Medium")
        high_n = dist.get("High")
        if all(isinstance(v, (int, float)) for v in (low_n, med_n, high_n)):
            story.append(create_bar_chart_drawing(
                "Existing Network: Risk Tier Distribution",
                [
                    ("Low Risk", int(low_n), BRAND_GREEN),
                    ("Medium Risk", int(med_n), BRAND_ORANGE),
                    ("High Risk", int(high_n), BRAND_RED),
                ],
            ))
            story.append(Spacer(1, 8))

        story.append(kv_table([
            ["Risk Classification Tier", "Site Count"],
            ["Low Risk (Performing Normally)", str(_fmt(dist.get("Low")))],
            ["Medium Risk (Monitor Performance)", str(_fmt(dist.get("Medium")))],
            ["High Risk (Underperformance Watch List)", str(_fmt(dist.get("High")))],
        ], colWidths=[260, 208]))

        top_zones = risk.get("top_10_zones") or {}
        if top_zones:
            story.append(Spacer(1, 12))
            story.append(Paragraph("Top Zones by Underperformance Risk Rate", h3_style))
            rows = [["Zone Name", "Risk Rate (%)", "Total Sites"]]
            for zone, v in top_zones.items():
                rate = v.get("risk_rate", 0)
                rate_str = f"{float(rate):.2f}%" if isinstance(rate, (int, float)) else str(rate)
                rows.append([zone, rate_str, str(v.get("count", 0))])
            story.append(kv_table(rows, colWidths=[240, 114, 114]))
            story.append(Spacer(1, 4))
            story.append(Paragraph(
                "<i>* Note: Zones with fewer than 10 sites are shown for completeness but carry higher statistical uncertainty due to small sample size.</i>",
                note_style
            ))

    def _draw_footer(canvas_obj, doc_obj):
        canvas_obj.saveState()
        canvas_obj.setFont("Helvetica", 8)
        canvas_obj.setFillColor(colors.HexColor("#94a3b8"))
        canvas_obj.drawString(16 * mm, 10 * mm, "OptiATM Network Report \u2014 Confidential, prepared for customer briefing")
        canvas_obj.drawRightString(A4[0] - 16 * mm, 10 * mm, f"Page {doc_obj.page}")
        canvas_obj.restoreState()

    doc.build(story, onFirstPage=_draw_footer, onLaterPages=_draw_footer)
    buffer.seek(0)
    return buffer


# ---------------------------------------------------------------------------
# Excel
# ---------------------------------------------------------------------------

def generate_excel_report(payload: dict) -> io.BytesIO:
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment
    from openpyxl.utils import get_column_letter

    HEADER_FILL = PatternFill(start_color="2563EB", end_color="2563EB", fill_type="solid")
    HEADER_FONT = Font(name="Calibri", bold=True, color="FFFFFF", size=10)
    TITLE_FONT = Font(name="Calibri", bold=True, size=14, color="2563EB")
    LABEL_FONT = Font(name="Calibri", bold=True, size=10)
    BODY_FONT = Font(name="Calibri", size=10)

    wb = Workbook()

    def write_table(ws, start_row, headers, rows, col_widths=None):
        for c_idx, header in enumerate(headers, start=1):
            cell = ws.cell(row=start_row, column=c_idx, value=header)
            cell.font = HEADER_FONT
            cell.fill = HEADER_FILL
            cell.alignment = Alignment(horizontal="left", vertical="center")
        for r_offset, row in enumerate(rows, start=1):
            for c_idx, value in enumerate(row, start=1):
                cell = ws.cell(row=start_row + r_offset, column=c_idx, value=value)
                cell.font = BODY_FONT
        if col_widths:
            for c_idx, width in enumerate(col_widths, start=1):
                ws.column_dimensions[get_column_letter(c_idx)].width = width
        return start_row + len(rows) + 2  # next free row

    # --- Overview sheet ------------------------------------------------
    ws = wb.active
    ws.title = "Overview"
    ws["A1"] = "OptiATM Network Report"
    ws["A1"].font = TITLE_FONT
    ws["A2"] = f"Generated {datetime.now().strftime('%d %b %Y, %H:%M')} - Hyderabad ATM Network"
    ws["A2"].font = Font(name="Calibri", size=9, italic=True, color="64748B")

    kpis = payload.get("kpis") or {}
    write_table(ws, 4, ["Metric", "Value"], [
        ["Active Own ATMs", _fmt(kpis.get("active_atms"))],
        ["Demographic Coverage", _fmt(kpis.get("coverage_pct"))],
        ["Total Daily Transaction Volume", _fmt(kpis.get("total_daily_tx"))],
        ["Avg Network Uptime (30-day)", _fmt(kpis.get("avg_uptime"))],
    ], col_widths=[36, 22])

    mm_data = payload.get("model_metrics")
    if mm_data:
        ws["A10"] = "ML Demand Prediction Model"
        ws["A10"].font = LABEL_FONT
        write_table(ws, 11, ["Metric", "Value"], [
            ["Active Model", _fmt(mm_data.get("best_model_name"))],
            ["R2 Score (Test)", _fmt(mm_data.get("r2"))],
            ["MAE (Transactions)", _fmt(mm_data.get("mae"))],
            ["RMSE (Transactions)", _fmt(mm_data.get("rmse"))],
            ["Prediction Accuracy", _fmt(mm_data.get("accuracy"))],
        ], col_widths=[36, 22])

    # --- Optimization sheet ---------------------------------------------
    opt = payload.get("optimization")
    if opt:
        ws_opt = wb.create_sheet("Optimization Results")
        method_labels = {"mclp": "Maximal Coverage (MCLP)", "p-median": "p-Median (Min Travel Distance)", "revenue": "ML-Revenue Maximizer"}
        ws_opt["A1"] = "Location Optimization Results"
        ws_opt["A1"].font = TITLE_FONT
        ws_opt["A2"] = f"Method: {method_labels.get(opt.get('method'), opt.get('method', '--'))}   |   Budget (k): {_fmt(opt.get('k'))}   |   Radius/Spacing: {_fmt(opt.get('radius'))} km"
        ws_opt["A2"].font = Font(name="Calibri", size=9, color="475569")
        if opt.get("summary"):
            ws_opt["A3"] = opt["summary"]
            ws_opt["A3"].font = Font(name="Calibri", size=9, italic=True, color="475569")

        m = opt.get("metrics") or {}
        next_row = write_table(ws_opt, 5, ["Impact Metric", "Before", "After", "Change"], [
            ["Population Coverage (%)", m.get("coverage_before"), m.get("coverage_after"), m.get("coverage_increase")],
            ["Avg Travel Distance (km)", m.get("avg_dist_before_km"), m.get("avg_dist_after_km"), m.get("avg_dist_reduction_km")],
        ], col_widths=[28, 14, 14, 14])

        selected = opt.get("selected_candidates") or []
        if selected:
            rows = []
            for c in selected:
                payback = c.get("payback_period")
                payback_val = payback if isinstance(payback, (int, float)) and payback > 0 else "Never"
                rows.append([
                    c.get("candidate_id", "--"), c.get("name", "--"), c.get("zone_name", "--"),
                    c.get("site_type", "--"), c.get("latitude"), c.get("longitude"),
                    c.get("rent_cost"), c.get("predicted_daily_transactions"),
                    c.get("roi_index"), payback_val,
                ])
            write_table(ws_opt, next_row, [
                "Candidate ID", "Name", "Zone", "Site Type", "Latitude", "Longitude",
                "Rent Cost (INR)", "Predicted Daily Tx", "ROI Index", "Payback (months)",
            ], rows, col_widths=[14, 22, 16, 14, 11, 11, 14, 16, 11, 14])

    # --- Removal simulation sheet -----------------------------------------
    rem = payload.get("removal_simulation")
    if rem:
        ws_rem = wb.create_sheet("Removal Simulation")
        ws_rem["A1"] = "ATM Removal Impact Simulation"
        ws_rem["A1"].font = TITLE_FONT
        ws_rem["A2"] = f"Site: {_fmt(rem.get('atm_id'))} ({_fmt(rem.get('zone_name'))}, {_fmt(rem.get('site_type'))})   |   Service Radius: {_fmt(rem.get('radius_km'))} km"
        ws_rem["A2"].font = Font(name="Calibri", size=9, color="475569")

        next_row = write_table(ws_rem, 4, ["Impact Metric", "Before Removal", "After Removal", "Change"], [
            ["Population Coverage (%)", rem.get("coverage_before"), rem.get("coverage_after"), -abs(rem.get("coverage_loss", 0))],
            ["Avg Travel Distance (km)", rem.get("avg_dist_before_km"), rem.get("avg_dist_after_km"), rem.get("avg_dist_increase_km")],
        ], col_widths=[28, 16, 16, 14])

        write_table(ws_rem, next_row, ["Risk Factor", "Value"], [
            ["Population Losing All Coverage", rem.get("affected_population")],
            ["Avg Daily Transactions (this site)", rem.get("avg_daily_transactions")],
            ["Monthly Revenue at Risk (INR)", rem.get("monthly_revenue_at_risk")],
            ["Monthly Rent Saved (INR)", rem.get("monthly_rent_saved")],
            ["Net Monthly Profit Impact (INR)", rem.get("net_monthly_profit_impact")],
        ], col_widths=[34, 20])

    # --- Risk analytics sheet --------------------------------------------
    risk = payload.get("risk_summary")
    if risk:
        ws_risk = wb.create_sheet("Risk Analytics")
        ws_risk["A1"] = "Risk Analytics Summary"
        ws_risk["A1"].font = TITLE_FONT
        dist = risk.get("risk_distribution") or {}
        next_row = write_table(ws_risk, 3, ["Risk Tier", "Site Count"], [
            ["Low", dist.get("Low")],
            ["Medium", dist.get("Medium")],
            ["High", dist.get("High")],
        ], col_widths=[20, 14])

        top_zones = risk.get("top_10_zones") or {}
        if top_zones:
            rows = []
            for zone, v in top_zones.items():
                rate = v.get("risk_rate", 0)
                rate_str = f"{float(rate):.2f}%" if isinstance(rate, (int, float)) else str(rate)
                rows.append([zone, rate_str, v.get("count", 0)])
            next_row = write_table(ws_risk, next_row, ["Zone", "Risk Rate (%)", "Site Count"], rows, col_widths=[24, 16, 14])
            ws_risk.cell(row=next_row - 1, column=1, value="* Note: Zones with fewer than 10 sites are shown for completeness but carry higher statistical uncertainty due to small sample size.")
            ws_risk.cell(row=next_row - 1, column=1).font = Font(name="Calibri", size=9, italic=True, color="64748B")

    buffer = io.BytesIO()
    wb.save(buffer)
    buffer.seek(0)
    return buffer